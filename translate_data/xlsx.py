from time import perf_counter
from openpyxl import workbook as wbook
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
#модуль для выгрузки данных в xlsx excell

class Xlsx(object):
    width_list = [2, 2, 2, 10, 9, 28, 28, 82, 10, 2, 2, 2, 2, 10]
    len_with_list = 0
    open_file = False

    def tables_null(self, index, row, column):
        for x in range(self.len_with_list - 1):
            self.ws.cell(index + row, column + x, ' ')

    def tables(self, index, row, column, column_list):
        colors = '00000000'
        bf = False
        for col in column_list:
            if not (col[2] is None or col[2]) is None:
                if col[2] == col[0]:
                    colors = col[1]
                    bf = True
                    break
        for col in range(len(column_list)):
            x = self.ws.cell(index + row, column + col + 1, column_list[col][0])
            x.font = Font(b=bf, color=colors)

    def write_file_xlsx(self, path_file, block):  # запись списка в фаил

        for index in range(len(block)):
            summa_op = lambda x, y: y if (x is True) else -y
            column_list = [
                [block[index].date, None, None],
                [block[index].number, None, None],
                [block[index].payer_name, None, None],
                [block[index].recipient_name, None, None],
                [block[index].payment_purpose, None, None],
                [summa_op(block[index].status, block[index].summa), None, None],
                [self.substitution(block[index].netting, True, '+', '-'), '00FF0000', '+'],
                [self.substitution(block[index].cash, True, '+', '-'), '000000FF', '+'],
                [self.substitution(block[index].tax, True, '+', '-'), '00008000', '+'],
                [self.substitution(block[index].settlement_bank, '40702810712300031547', 'A', 'T'), None, None],
                [self.substitution(block[index].cash_register, True, 'K', ''), '00800080', 'K']]

            self.len_with_list = len(column_list)
            if index == 0 and self.open_file is False: # если первая ячейка или фаил не существовал установить ширины ячеек
                for x in range(0, self.len_with_list + 1, self.len_with_list):
                    for column in range(1, self.len_with_list + 1):
                        self.ws.column_dimensions[get_column_letter(column + x)].width = self.width_list[column - 1]

            if block[index].owner == 'B':
                self.tables(index, 2, 3, column_list)
                self.tables_null(index, 2, self.len_with_list + 4)
            else:
                self.tables(index, 2, self.len_with_list + 3, column_list)
                self.tables_null(index, 2, 4)
        try:
            start_time = perf_counter()
            self.wb.save(path_file)
            print(perf_counter() - start_time)  # расчет времени
        except ():
            print('----Возжно уже открыт данный фаил----')

    def __init__(self, path_file_xlsx, block):
        try:
            self.wb =load_workbook(filename=path_file_xlsx)
            self.ws= self.wb['Sheet']
            self.open_file = True
        except:
            self.wb = wbook.Workbook()
            self.ws = self.wb.active
        self.len_with_list = len(self.width_list)
        self.write_file_xlsx(path_file_xlsx, block)


    @staticmethod
    def substitution(value, value_find, value_True, value_false) -> str:
        if value is not None or value_find is not None:
            if isinstance(value, str) is True and isinstance(value_find, str) is True:
                if value == value_find:
                    return value_True
                else:
                    return value_false

            if isinstance(value, bool) is True and isinstance(value_find, bool) is True:
                if value is value_find:
                    return value_True
                else:
                    return value_false
        return 'N'
